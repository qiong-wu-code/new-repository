#!/usr/bin/env python3
"""
Collect regression results for all (or specified) models and write an xlsx report.

Reads artifacts only (Converter_result/), does not run any conversion.

Usage:
  collect_results.py --models-root ./models --output-xlsx results/report.xlsx
  collect_results.py --models-root ./models --only m001,m003 --output-xlsx ...
  collect_results.py --models-root ./models --order-file order.yaml --output-xlsx ...
"""
import argparse
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from parse_snr import parse_org_vs_hw
from extract_errors import find_log_file, extract_from_log

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({
        'status': 'error', 'error_kind': 'missing_dependency',
        'message': 'openpyxl required. Install: pip install openpyxl'
    }))
    sys.exit(3)

try:
    import yaml
except ImportError:
    yaml = None

EXIT_OK = 0
EXIT_USER_ERROR = 2
EXIT_RUNTIME_ERROR = 3
EXIT_PARTIAL = 4

logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s', stream=sys.stderr)
log = logging.getLogger(__name__)

MODEL_DIR_PATTERN = re.compile(r'^(m\d+)_(.+)$')
LOW_SNR_THRESHOLD = 25.0

FILL_OK = PatternFill('solid', fgColor='C6EFCE')
FILL_FAILED = PatternFill('solid', fgColor='FFC7CE')
FILL_NOT_RUN = PatternFill('solid', fgColor='D9D9D9')
FILL_LOW_SNR = PatternFill('solid', fgColor='FFEB9C')
FONT_HEADER = Font(bold=True)
ALIGN_WRAP = Alignment(wrap_text=True, vertical='top')


def parse_args():
    p = argparse.ArgumentParser(
        prog='collect_results',
        description='Batch-collect regression results, write xlsx report.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --models-root ./models --output-xlsx results/report.xlsx
  %(prog)s --models-root ./models --only m001,m003 --output-xlsx r.xlsx
  %(prog)s --models-root ./models --order-file order.yaml --output-xlsx r.xlsx
  %(prog)s --models-root ./models --output-xlsx r.xlsx --dry-run

Exit codes:
  0 = all ok
  2 = user error (bad args)
  3 = runtime error
  4 = partial success (some models had collect errors)
""",
    )
    p.add_argument('--models-root', required=True, type=Path, metavar='DIR')
    p.add_argument('--output-xlsx', required=True, type=Path, metavar='PATH')
    p.add_argument('--only', default=None, metavar='IDS')
    p.add_argument('--order-file', default=None, type=Path, metavar='PATH')
    p.add_argument('--low-snr-threshold', type=float, default=LOW_SNR_THRESHOLD, metavar='N')
    p.add_argument('--dry-run', action='store_true')
    p.add_argument('--force', action='store_true')
    p.add_argument('-v', '--verbose', action='store_true')
    return p.parse_args()


def emit(payload, exit_code=EXIT_OK):
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.exit(exit_code)


def discover_models(root: Path, only_ids=None):
    models = []
    for d in sorted(root.iterdir()):
        if not d.is_dir():
            continue
        m = MODEL_DIR_PATTERN.match(d.name)
        if not m:
            continue
        mid, mname = m.group(1), m.group(2)
        if only_ids and mid not in only_ids:
            continue
        models.append({'id': mid, 'name': mname, 'path': d})
    return models


def load_order(order_file: Path):
    if yaml is None:
        raise RuntimeError('PyYAML required for --order-file')
    data = yaml.safe_load(order_file.read_text()) or []
    if not isinstance(data, list):
        raise ValueError(f'--order-file must contain a YAML list')
    return [str(x) for x in data]


def apply_order(models, order):
    by_id = {m['id']: m for m in models}
    ordered, seen = [], set()
    for mid in order:
        if mid in by_id and mid not in seen:
            ordered.append(by_id[mid])
            seen.add(mid)
    for m in models:
        if m['id'] not in seen:
            ordered.append(m)
    return ordered


def format_snr_display(snr) -> str:
    """Render snr for the xlsx 'SNR' column: float or list -> string."""
    if snr is None:
        return ''
    if isinstance(snr, list):
        return '[' + ', '.join(f'{v:.2f}' for v in snr) + ']'
    return f'{snr:.2f}'


def collect_one(model: dict) -> dict:
    """Collect a row dict for a single model."""
    cr = model['path'] / 'Converter_result'
    row = {
        'model_id': model['id'],
        'model_name': model['name'],
        'dtype_org': '',
        'dtype_hw': '',
        'status': '',
        'snr': None,
        'snr_min': None,
        'is_multi_output': False,
        'sdk_version': '',
        'error_summary': '',
    }

    if not cr.is_dir():
        row['status'] = 'not_run'
        row['error_summary'] = 'no Converter_result/'
        return row

    # Try SNR first
    snr_result = parse_org_vs_hw(cr / 'org_vs_hw.txt')
    if snr_result.get('status') == 'ok':
        row['status'] = 'ok'
        row['snr'] = snr_result.get('snr')
        row['snr_min'] = snr_result.get('snr_min')
        row['is_multi_output'] = snr_result.get('is_multi_output', False)
        row['sdk_version'] = snr_result.get('sdk_version', '')
        row['dtype_org'] = snr_result.get('dtype_org', '')
        row['dtype_hw'] = snr_result.get('dtype_hw', '')
        return row

    # Failed: look for error in log
    row['status'] = 'failed'
    log_file = find_log_file(cr)
    if log_file:
        err = extract_from_log(log_file, max_chars=500, tail_mb=10)
        if err.get('status') == 'found':
            row['error_summary'] = err.get('snippet', '')
        else:
            row['error_summary'] = f"log exists but no error pattern matched: {log_file.name}"
    else:
        if snr_result.get('status') == 'missing':
            row['error_summary'] = 'no org_vs_hw.txt and no log file'
        else:
            row['error_summary'] = f"snr parse: {snr_result.get('status')}; no log"

    # Even for failed case, keep partial info if parse_snr got some fields
    row['sdk_version'] = snr_result.get('sdk_version', '')
    row['dtype_org'] = snr_result.get('dtype_org', '')
    row['dtype_hw'] = snr_result.get('dtype_hw', '')
    return row


def write_xlsx(rows, output: Path, low_snr_threshold: float):
    wb = Workbook()
    ws = wb.active
    ws.title = 'SNR Results'

    headers = ['#', 'Model ID', 'Model Name', 'Dtype (org→hw)',
               'Status', 'SNR', 'SNR min', 'SDK Ver', 'Error']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = FONT_HEADER

    for idx, row in enumerate(rows, start=1):
        dtype_combined = ''
        if row['dtype_org'] or row['dtype_hw']:
            dtype_combined = f"{row['dtype_org']} → {row['dtype_hw']}"

        ws.append([
            idx,
            row['model_id'],
            row['model_name'],
            dtype_combined,
            row['status'],
            format_snr_display(row['snr']),
            row['snr_min'] if row['snr_min'] is not None else '',
            row['sdk_version'],
            row['error_summary'],
        ])
        excel_row = idx + 1

        # Status color
        status_cell = ws.cell(row=excel_row, column=5)
        if row['status'] == 'ok':
            status_cell.fill = FILL_OK
        elif row['status'] == 'failed':
            status_cell.fill = FILL_FAILED
        elif row['status'] == 'not_run':
            status_cell.fill = FILL_NOT_RUN

        # Low SNR highlight: check snr_min uniformly (works for multi-output)
        if row['snr_min'] is not None and row['snr_min'] < low_snr_threshold:
            ws.cell(row=excel_row, column=6).fill = FILL_LOW_SNR  # SNR col
            ws.cell(row=excel_row, column=7).fill = FILL_LOW_SNR  # SNR min col

        ws.cell(row=excel_row, column=9).alignment = ALIGN_WRAP

    widths = [5, 10, 26, 16, 10, 20, 10, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    args = parse_args()
    if args.verbose:
        log.setLevel(logging.DEBUG)

    errors = []
    if not args.models_root.is_dir():
        errors.append(f'--models-root does not exist: {args.models_root}')
    if args.output_xlsx.exists() and not args.force and not args.dry_run:
        errors.append(f'--output-xlsx exists (use --force to overwrite): {args.output_xlsx}')
    if args.order_file and not args.order_file.is_file():
        errors.append(f'--order-file does not exist: {args.order_file}')
    if errors:
        emit({'status': 'error', 'error_kind': 'invalid_args', 'errors': errors},
             EXIT_USER_ERROR)

    only_ids = set(args.only.split(',')) if args.only else None
    models = discover_models(args.models_root, only_ids)
    if not models:
        emit({'status': 'error', 'error_kind': 'no_models_found',
              'models_root': str(args.models_root), 'only_filter': args.only},
             EXIT_USER_ERROR)

    if args.order_file:
        try:
            order = load_order(args.order_file)
            models = apply_order(models, order)
        except Exception as e:
            emit({'status': 'error', 'error_kind': 'order_file_error', 'message': str(e)},
                 EXIT_USER_ERROR)

    log.info(f'Processing {len(models)} models')

    if args.dry_run:
        emit({
            'status': 'dry_run',
            'would_process': [{'id': m['id'], 'name': m['name']} for m in models],
            'total': len(models),
            'output_xlsx': str(args.output_xlsx),
        }, EXIT_OK)

    rows = []
    summary = {'ok': 0, 'failed': 0, 'not_run': 0, 'collect_error': 0}
    low_snr_models = []
    multi_output_models = []
    collect_errors = []

    for m in models:
        try:
            row = collect_one(m)
            rows.append(row)
            summary[row['status']] = summary.get(row['status'], 0) + 1
            if row['snr_min'] is not None and row['snr_min'] < args.low_snr_threshold:
                low_snr_models.append(row['model_id'])
            if row['is_multi_output']:
                multi_output_models.append(row['model_id'])
        except Exception as e:
            log.exception(f'Collect failed: {m["id"]}')
            summary['collect_error'] += 1
            collect_errors.append({'id': m['id'], 'error': str(e)})
            rows.append({
                'model_id': m['id'], 'model_name': m['name'],
                'dtype_org': '', 'dtype_hw': '',
                'status': 'collect_error', 'snr': None, 'snr_min': None,
                'is_multi_output': False, 'sdk_version': '',
                'error_summary': f'collect_results exception: {e}',
            })

    try:
        write_xlsx(rows, args.output_xlsx, args.low_snr_threshold)
    except Exception as e:
        log.exception('xlsx write failed')
        emit({'status': 'error', 'error_kind': 'xlsx_write_failed', 'message': str(e)},
             EXIT_RUNTIME_ERROR)

    exit_code = EXIT_PARTIAL if summary['collect_error'] > 0 else EXIT_OK
    emit({
        'status': 'ok' if exit_code == EXIT_OK else 'partial',
        'total': len(models),
        'summary': summary,
        'low_snr_count': len(low_snr_models),
        'low_snr_models': low_snr_models,
        'multi_output_count': len(multi_output_models),
        'multi_output_models': multi_output_models,
        'collect_errors': collect_errors,
        'output_xlsx': str(args.output_xlsx),
        'timestamp': datetime.now().isoformat(),
    }, exit_code)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
    except SystemExit:
        raise
    except Exception as e:
        log.exception('Unhandled exception')
        emit({'status': 'error', 'error_kind': 'unhandled_exception', 'message': str(e)},
             EXIT_RUNTIME_ERROR)
